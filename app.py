from flask import Flask
from flask import render_template, send_from_directory,url_for, request
from openpyxl import load_workbook
import json
import os

UPLOAD_FOLDER = '/path/to/the/uploads'
ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'])
root_file_folder ='app/public/local'
app = Flask(__name__)
# app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/', methods = ['get'])
def home():
  return render_template('index.html')

@app.route('/get/excel', methods = ['get'])
@app.route('/upload/excel', methods = ['post'])
def upload():
  # file = request.files['file']
  book = load_workbook('i18n.xlsx')
  res = {}
  for n, name in enumerate(book.sheetnames):
    res[name] = {}
    sheet = book.get_sheet_by_name(name)
    for i, r in enumerate(sheet.rows):
      res[name][i] = {}
      for k, c in enumerate(r):
        res[name][i][k] = c.value
  return json.dumps(res)
    
if __name__ == '__main__':
  app.run()